import pyodbc
import pandas as pd
from openpyxl import load_workbook
import tempfile
import os
import mysql.connector

def leggi_opzioni_da_file(nome_file):
    percorso_file=f"VAR_DINAMICHE\\{nome_file}"
    with open(percorso_file, 'r', encoding='utf-8') as file:
        contenuto = file.read()
        opzioni = [opzione.strip() for opzione in contenuto.split(',') if opzione.strip()]
    return opzioni


def Ask_CF(CFS):
    config = {
        'user': 'root',
        'host': 'localhost',
        'database': 'RNAuser'
    }

    base_query = "SELECT * FROM `aiuti_individuali` WHERE 1=1"
    query_params = []
    
    if len(CFS) == 1:
        base_query += " AND `C.F. Beneficiario` = %s"
        query_params.append(CFS['CF_1'].strip())
    elif len(CFS) > 1:
        placeholders = ', '.join(['%s'] * len(CFS))
        base_query += f" AND `C.F. Beneficiario` IN ({placeholders})"
        for key in CFS:
            query_params.append(CFS[key].strip())
    else:
        return None

    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()

    try:
        cursor.execute(base_query, query_params)
        rows = cursor.fetchall()

        columns = [column[0] for column in cursor.description]

        df = pd.DataFrame.from_records(rows, columns=columns)
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()

    return df

def CREA_DF(AMBITO):
    
    if not AMBITO.empty: 
        selected_columns = AMBITO#.iloc[:, [2, 3, 4, 8, 10]]
        new_df = pd.DataFrame({
            'DEN' : selected_columns.iloc[:, 10],
            'CFB' : selected_columns.iloc[:, 11],
            'DATA' : selected_columns.iloc[:, 7],
            'autorita' : selected_columns.iloc[:, 14],
            'Tit_misura' : selected_columns.iloc[:, 1],       
            'Norma_misura' : selected_columns.iloc[:, 3],
            'AIUTO' : selected_columns.iloc[:, 23]
        })
        return new_df

def COMPILA_EXCEL_agr(ambito):

    excel_template = "C:\\Users\\g.stella\\Desktop\\CIAO\\DE MINIMIS 1408 - AGRICOLTURA.xlsx"
    wb = load_workbook(excel_template)
    ws = wb.active

    mapping = {
    'C': 'DEN',
    'D': 'CFB',
    'E': 'DATA',
    'G': 'autorita',
    'H': 'Tit_misura',
    'I': 'Norma_misura',
    'J': 'AIUTO'
    }

    start_row = 6
    for index, row in ambito.iterrows():
        for excel_col, df_col in mapping.items():
            ws[f'{excel_col}{start_row + index}'] = row[df_col]

    with tempfile.TemporaryDirectory() as tmpdirname:
        doc_path = os.path.join(tmpdirname, "TENTATIVO.xlsx")
        wb.save(doc_path)
        with open(doc_path, "rb") as f:
            doc_contents = f.read()
        return doc_contents

def COMPILA_EXCEL_pes(ambito):

    excel_template = "C:\\Users\\g.stella\\Desktop\\CIAO\\DE MINIMIS 717 - PESCA E ACQUACOLTURA.xlsx"
    wb = load_workbook(excel_template)
    ws = wb.active

    mapping = {
    'C': 'DEN',
    'D': 'CFB',
    'E': 'DATA',
    'G': 'autorita',
    'H': 'Tit_misura',
    'I': 'Norma_misura',
    'J': 'AIUTO'
    }

    start_row = 6
    for index, row in ambito.iterrows():
        for excel_col, df_col in mapping.items():
            ws[f'{excel_col}{start_row + index}'] = row[df_col]

    with tempfile.TemporaryDirectory() as tmpdirname:
        doc_path = os.path.join(tmpdirname, "TENTATIVO.xlsx")
        wb.save(doc_path)
        with open(doc_path, "rb") as f:
            doc_contents = f.read()
        return doc_contents

def COMPILA_EXCEL_gen(ambito):

    excel_template = "C:\\Users\\g.stella\\Desktop\\CIAO\\DE MINIMIS 2831 - GENERALE.xlsx"
    wb = load_workbook(excel_template)
    ws = wb.active

    mapping = {
    'C': 'DEN',
    'D': 'CFB',
    'E': 'DATA',
    'H': 'autorita',
    'I': 'Tit_misura',
    'J': 'Norma_misura',
    'K': 'AIUTO'
    }
    try:
        if ambito != None:
            pass
    except:
        start_row = 6
        for index, row in ambito.iterrows():
            for excel_col, df_col in mapping.items():
                ws[f'{excel_col}{start_row + index}'] = row[df_col]

    with tempfile.TemporaryDirectory() as tmpdirname:
        doc_path = os.path.join(tmpdirname, "TENTATIVO.xlsx")
        wb.save(doc_path)
        with open(doc_path, "rb") as f:
            doc_contents = f.read()
        return doc_contents

def ricerca_avanzata(parametro): 
    config = {
        'user': 'root',
        'host': 'localhost',
        'database': 'RNAuser'
    }

    base = "SELECT * FROM `aiuti_individuali` WHERE 1=1 "
    ricerca = []
    for key in parametro:
        if parametro[key] != "":
            if key == "Data_Start":
                base += f"AND `DATA CONCESSIONE` > %s "
            elif key == "Data_End":
                base += f"AND `DATA CONCESSIONE` < %s "
            else:
                base += f"AND `{key}` LIKE %s "
            ricerca.append(parametro[key])
    base += f"ORDER BY `DATA CONCESSIONE` DESC LIMIT %s"
    ricerca.append(100000)
    if base != "SELECT * FROM `aiuti_individuali` WHERE 1=1 ORDER BY `DATA CONCESSIONE` DESC LIMIT %s":
        try:
            connection = mysql.connector.connect(**config)
            cursor = connection.cursor(dictionary=True)
            
            cursor.execute(base, ricerca)
            result = cursor.fetchall()
            columns = [column[0] for column in cursor.description]
            df = pd.DataFrame.from_records(result, columns=columns)
            return df
        except mysql.connector.Error as err:
            print(f"Errore: {err}")
            return None
        finally:
            if connection.is_connected():
                cursor.close()
                connection.close()
    else: 
        pass
        
def Excel_avanzato(valori):
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        valori.to_excel(tmp_file.name, index=False, engine='openpyxl')
        
        temp_file_path = tmp_file.name
        with open(temp_file_path, 'rb') as f:
            excel_data = f.read()

        return excel_data